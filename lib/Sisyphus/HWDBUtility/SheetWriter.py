#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
bin/advanced-upload-docket.py
Copyright (c) 2023 Regents of the University of Minnesota
Author: Alex Wagner <wagn0033@umn.edu>, Dept. of Physics and Astronomy
"""

from Sisyphus.Configuration import config
logger = config.getLogger(__name__)

from Sisyphus.Utils import Fonts

import sys
import argparse
import json, json5
from copy import copy

import Sisyphus.RestApiV1 as ra
from Sisyphus.RestApiV1 import Utilities as ut

import pandas as pd
import openpyxl.styles as styles
import openpyxl.styles.borders as borders
from openpyxl.utils import get_column_letter

#from Sisyphus.HWDBUploader import Docket

from Sisyphus.Utils.Terminal.Style import Style

class ExcelWriter:
    #default_font = "Arial"
    #default_font = "DejaVu Sans"
    #default_font = "DejaVu Sans Condensed"
    #default_font = "Liberation Sans"
    default_normal_font = "Calibri"
    default_header_font = "Calibri Bold"
    default_font_size = 10

    def __init__(self, filename):
        #{{{
        self.filename = filename
        self.writer = pd.ExcelWriter(self.filename, engine="openpyxl")
        
        wb = self.writer.book

        # Create style "key"
        key_style = styles.NamedStyle(name="key")

        key_style.alignment = styles.Alignment(
                                    horizontal='left',
                                    vertical='top',
                                    wrap_text=False)
        key_style.font = styles.Font(
                            name=self.default_normal_font,
                            size=self.default_font_size,
                            bold=True)
        wb.add_named_style(key_style)
        self.key_style_ImageFont = Fonts.get_font(self.default_header_font, self.default_font_size)


        # Create style "data"
        data_style = styles.NamedStyle(name="data")
        data_style.alignment = styles.Alignment(
                                    #horizontal='left',
                                    vertical='top',
                                    wrap_text=True)
        data_style.font = styles.Font(
                            name=self.default_normal_font,
                            size=self.default_font_size,
                            bold=False)
        wb.add_named_style(data_style)
        self.data_style_ImageFont = Fonts.get_font(self.default_normal_font, self.default_font_size)
        
        # Create style "header_data"
        header_data_style = styles.NamedStyle(name="header_data")
        header_data_style.alignment = styles.Alignment(
                                    horizontal='left',
                                    vertical='top',
                                    wrap_text=True)
        header_data_style.font = styles.Font(
                            name=self.default_normal_font,
                            size=self.default_font_size,
                            bold=False)
        wb.add_named_style(header_data_style)
        self.header_data_style_ImageFont = Fonts.get_font(self.default_normal_font, self.default_font_size)

        # Borders don't seem to be working, but I don't know if it's just
        # because I'm opening these in LibreOffice Calc instead of Excel.
        side_1 = borders.Side(color="000000", border_style='thin')
        self.thin_border = borders.Border(
                    left=side_1, 
                    right=side_1, 
                    top=side_1, 
                    bottom=side_1)


        #}}}

    def close(self):
        #{{{
        self.writer.close()
        #}}}

    def add_sheet(self, sheetname, header=None, table=None):
        #{{{
        writer = self.writer
        header_startrow = 0
        table_startrow = 0

        if header is not None:
            # Assume the header is in key/value pairs.
            # What we need to do to make it an acceptable dataframe is to make
            # the value into a list containing that (single) value.
            df_header = pd.DataFrame({k:[v] for k,v in header.items()}).transpose()
            df_header.to_excel(
                    writer,
                    sheet_name=sheetname,
                    startrow=header_startrow,
                    index=True,
                    header=False)
            table_startrow = len(header) + 1

        if table is not None:
            df_table = pd.DataFrame(table)
            df_table.to_excel(
                    writer,
                    sheet_name=sheetname,
                    startrow=table_startrow,
                    index=False,
                    header=True)

        # Now that we've written the data, let's do some formatting so the 
        # spreadsheet doesn't look like 10 lbs. of manure in a 5 lb. bag.
        ws = writer.sheets[sheetname]

        clamp = lambda v, L, H: max(min(v, H), L)

        def set_col_width(ws, colnum, chars=9):
            padding = 0.75
            width_per_char = 1.05
            min_chars = 10
            max_chars = 50
            width = padding + clamp(chars, min_chars, max_chars) * width_per_char
            ws.column_dimensions[get_column_letter(colnum)].width = width
            return width

        def set_row_height(ws, rownum, lines=1):
            padding = 3
            height_per_line = 10
            min_lines = 1
            max_lines = 15
            height = padding + clamp(lines, min_lines, max_lines) * height_per_line
            ws.row_dimensions[rownum].height = height
            return height

        def OLD_calc_ideal_size(value, max_width=65):
            # TODO: delete when I'm sure that the ImageFont method works.

            # if it weren't for line feeds or text wrapping, this would
            # just be the width of the value as a string
            # The estimate below isn't great, but it'll do for now.
            parts = str(value).split("\n")
            chars = max([len(part) for part in parts])
            lines = sum([1 + len(part)//max_width for part in parts])

            return (chars, lines)

        def calc_ideal_size(imageFont, value, max_width=65):
            # if it weren't for line feeds or text wrapping, this would
            # just be the width of the value as a string
            # The estimate below isn't great, but it'll do for now.
            parts = str(value).split("\n")

            # TODO: I sort of shoehorned the ImageFont method of getting the 
            # length, and didn't have a chance to figure out what the actual
            # conversion is from pixels to "character width", so I just used
            # what seemed to work.
            chars = 0.19 * max([  imageFont.getlength(part) for part in parts])


            lines = sum([1 + len(part)//max_width for part in parts])

            return (chars, lines)
        
        # If the total width of the columns used isn't some minimum value,
        # add an extra column and make its width enough to get to the minimum.
        # (The extra column is necessary to not mess up the horizontal 
        # alignment of the last column.)
        min_sheet_width = 80

        cell_size = {}

        #for rownum in range(1, len(header)+1):
        #    ws.cell(rownum, 1).alignment = key_alignment

        # Calculate the "best fit" widths for columns.
        for row in ws.iter_rows():
            rownum = row[0].row
            
            for cell in row:
                r, c = cell.row, cell.column
                cell.border = self.thin_border
        
                # We have different rules for the 'header' rows and 'table'
                # rows. The "values" in the header rows can influence the
                # row height, but not the column widths.
                if rownum < table_startrow:
                    if c == 1:
                        # header key
                        #w, h = OLD_calc_ideal_size(cell.value)
                        w, h = calc_ideal_size(self.key_style_ImageFont, cell.value)
                        cell_size[r, c] = (w, h)
                        cell.style = "key"
                    else:
                        # header value
                        #w, h = OLD_calc_ideal_size(cell.value, max_width=min_sheet_width)
                        w, h = calc_ideal_size(self.header_data_style_ImageFont,
                                    cell.value, max_width=min_sheet_width)
                        cell_size[r, c] = (0, h)
                        cell.style = "header_data"
                elif rownum == table_startrow + 1:
                    # table key
                    #w, h = OLD_calc_ideal_size(cell.value)
                    w, h = calc_ideal_size(self.key_style_ImageFont, cell.value)
                    cell_size[r, c] = (w, h)
                    cell.style = "key"
                else:
                    # table value
                    #w, h = OLD_calc_ideal_size(cell.value)
                    w, h = calc_ideal_size(self.data_style_ImageFont, cell.value)
                    cell_size[r, c] = (w, h)
                    cell.style = "data"

        max_row_size = {}
        max_col_size = {}
        for (r, c), (w, h) in cell_size.items():
            max_col_size[c] = max(max_col_size.get(c, 0), w)
            max_row_size[r] = max(max_row_size.get(r, 0), h)

        total_width, total_height = 0, 0
        for c, w in max_col_size.items():
            total_width += set_col_width(ws, c, w)
        for r, h in max_row_size.items():
            total_height += set_row_height(ws, r, h)

        
        if header is not None:
            if table is None:
                if total_width < min_sheet_width:
                    ws.column_dimensions[get_column_letter(2)].width = \
                                min_sheet_width - ws.column_dimensions[get_column_letter(1)].width
            else:
                end_column = ws.max_column
                if total_width < min_sheet_width:
                    ws.column_dimensions[get_column_letter(ws.max_column+1)].width = \
                                                            min_sheet_width - total_width
                    end_column = ws.max_column + 1
        
                # Merge the header rows 
                for rownum in range(1, table_startrow):
                    ws.merge_cells(
                                start_row=rownum, 
                                start_column=2, 
                                end_row=rownum, 
                                end_column=end_column)
                # Merge the blank row
                ws.merge_cells(
                            start_row=table_startrow, 
                            start_column=1, 
                            end_row=table_startrow, 
                            end_column=end_column)
        #}}}

def main(argv):
    pass

if __name__ == '__main__':
    sys.exit(main(sys.argv))


